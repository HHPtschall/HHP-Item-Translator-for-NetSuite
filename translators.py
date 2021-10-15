from openpyxl import load_workbook, Workbook
from datetime import datetime
from pytz import timezone


# Grabs the header values and puts them in the output file
def get_headers(input_sheet, output_sheet, header_length):
    for row in input_sheet.iter_rows(max_row=1,
                                     max_col=header_length,
                                     values_only=True):
        row = list(row)
        output_sheet.append(row)
        return output_sheet


# Loads in the document
def load_doc(input_file, sheet_name):
    file = input_file + ".xlsx"
    workbook = load_workbook(read_only=True, filename='Input Files/' + file)
    input_sheet = workbook[sheet_name]
    print(input_file, 'Loaded')
    return input_sheet


# Creates a new output workbook and worksheet
def create_outputs():
    output = Workbook()
    output_sheet = output.active
    return output, output_sheet


# Concatenates all input data into usable entries for output
def concat_data(input_sheet, entries):
    for row in input_sheet.iter_rows(min_row=2, max_col=11, values_only=True):
        # Converts the tuple row into a usable list and all its entries into strings
        row = list(row)
        for i in range(len(row)):
            row[i] = '' if row[i] == None else str(row[i])

        # Gets the specific P# data from entries
        temp = row[0]
        old_stuff = entries.get(temp)

        # Makes sure that we are getting dictionary values into a list
        if type(old_stuff) == list:
            old_values = old_stuff
        else:
            old_values = list(old_stuff.values())

        # Creates or adds to the initial entries
        for i in range(len(old_values) - 1):
            if old_values[i] == '':
                old_values[i] = row[i + 1]
            elif row[i + 1] not in old_values[i]:
                old_values[i] += '|' + row[i + 1]

        # Saves the new values to the existing dictionary
        entries[temp] = old_values
    return entries


# Adds each entry to the output file and saves the final output file
def append_and_save(entries, output, output_sheet, filename):
    for entry in entries:
        row_to_add = entries[entry]

        # Makes sure the entries are in a list format
        if type(row_to_add) is dict:
            row_to_add = list(row_to_add.values())

        # Adds the P# to the start of the list
        row_to_add.insert(0, entry)

        # Appends the full row to the file
        output_sheet.append(row_to_add)

    dt = ((timezone("US/Eastern")).localize(datetime.now())).strftime('%m.%d ')
    output.save(filename="Output Files/" + dt + filename + '.csv')
    print('File Created')


def heavy_translator(input_file):
    input_sheet = load_doc(input_file,
                           "Heavy Duty")  #Possibly need to make this modular

    output, output_sheet = create_outputs()
    output_sheet = get_headers(input_sheet, output_sheet, 7)

    # This is the dictionary of entries that will be put into the output
    entries = {}
    # Goes through the input file and puts data into the entries dictionary
    for row in input_sheet.iter_rows(min_row=2, max_col=11, values_only=True):
        # Converts the tuple row into a usable list and all its entries into strings
        row = list(row)
        for i in range(len(row)):
            row[i] = '' if row[i] == None else str(row[i])

        # The P# and key to the entries dictionary
        temp = row[0]

        # Adds in an entry for the first time
        if temp not in entries.keys():
            entries[temp] = {
                'Truck OEM': row[1],
                'Truck Engine Model': row[7],
                'Truck ESN/CPL/Liter': row[8],
                'Product Group': row[9],
                'Heavy String': row[10],
                'Internal ID': row[6]
            }

    entries = concat_data(input_sheet, entries)

    append_and_save(entries, output, output_sheet, 'Heavy Duty Output')


def mid_translator(input_file):
    input_sheet = load_doc(input_file,
                           "Midrange")  #Possibly need to make this modular

    output, output_sheet = create_outputs()
    output_sheet = get_headers(input_sheet, output_sheet, 6)

    # This is the dictionary of entries that will be put into the output
    entries = {}
    # Goes through the input file and puts data into the entries dictionary
    for row in input_sheet.iter_rows(min_row=2, max_col=11, values_only=True):
        # Converts the tuple row into a usable list and all its entries into strings
        row = list(row)
        for i in range(len(row)):
            row[i] = '' if row[i] == None else str(row[i])

        # The P# and key to the entries dictionary
        temp = row[0]

        # Adds in an entry for the first time
        if temp not in entries.keys():
            entries[temp] = {
                'Make': row[6],
                'Model': row[7],
                'Year': row[8],
                'Mid String': row[9],
                'Internal ID': row[5]
            }

    entries = concat_data(input_sheet, entries)

    append_and_save(entries, output, output_sheet, 'Mid Range Output')


# TODO Make the thing for this guy
def ag_translator(input_file):
    pass
